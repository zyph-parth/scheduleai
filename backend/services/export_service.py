"""
Export service for PDF and Excel timetable exports.
"""

import io
from typing import Any, Dict, List, Optional


def _slot_section_bindings(slot: Dict[str, Any]) -> List[tuple[int, str]]:
    section_ids = slot.get("section_ids") or ([slot["section_id"]] if slot.get("section_id") else [])
    section_labels = slot.get("section_labels") or []
    bindings: List[tuple[int, str]] = []

    for index, section_id in enumerate(section_ids):
        label = (
            section_labels[index]
            if index < len(section_labels)
            else slot.get("section_name") if len(section_ids) == 1
            else f"Section {section_id}"
        )
        bindings.append((int(section_id), str(label)))
    return bindings


def _section_entries(slots: List[Dict[str, Any]]) -> List[tuple[Optional[int], str]]:
    entries = sorted(
        {
            binding
            for slot in slots
            for binding in _slot_section_bindings(slot)
        },
        key=lambda item: (item[1], item[0]),
    )
    return entries or [(None, "Timetable")]


def _slots_for_section(slots: List[Dict[str, Any]], section_id: Optional[int]) -> List[Dict[str, Any]]:
    if section_id is None:
        return list(slots)
    return [
        slot
        for slot in slots
        if any(binding_id == section_id for binding_id, _ in _slot_section_bindings(slot))
    ]


def export_excel(
    slots: List[Dict[str, Any]],
    institution_name: str,
    periods_per_day: Dict,
    days_names: Optional[List[str]] = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    if days_names is None:
        days_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

    wb = Workbook()
    section_entries = _section_entries(slots)

    header_fill = PatternFill("solid", fgColor="1E293B")
    day_fill = PatternFill("solid", fgColor="334155")
    theory_fill = PatternFill("solid", fgColor="DBEAFE")
    lab_fill = PatternFill("solid", fgColor="D1FAE5")
    break_fill = PatternFill("solid", fgColor="FEF3C7")
    empty_fill = PatternFill("solid", fgColor="F8FAFC")
    border_side = Side(style="thin", color="CBD5E1")
    cell_border = Border(
        left=border_side,
        right=border_side,
        top=border_side,
        bottom=border_side,
    )

    def apply_header(cell, value, bold: bool = True, fg: str = "FFFFFF"):
        cell.value = value
        cell.font = Font(bold=bold, color=fg, name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border

    for section_id, section_name in section_entries:
        section_slots = _slots_for_section(slots, section_id)

        ws = wb.create_sheet(title=section_name[:31])
        ws.sheet_view.showGridLines = False

        max_col = max((max(periods) for periods in periods_per_day.values() if periods), default=7) + 3
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"{institution_name} - {section_name} Timetable"
        title_cell.fill = header_fill
        apply_header(title_cell, title_cell.value)
        ws.row_dimensions[1].height = 24

        working_days = sorted(periods_per_day.keys())
        all_periods = sorted({period for periods in periods_per_day.values() for period in periods})

        ws.cell(row=2, column=1).fill = header_fill
        apply_header(ws.cell(row=2, column=1), "Day / Period")

        for col_index, period in enumerate(all_periods, start=2):
            cell = ws.cell(row=2, column=col_index)
            cell.fill = header_fill
            apply_header(cell, f"P{period + 1}", fg="FFFFFF")

        ws.row_dimensions[2].height = 20

        for row_index, day in enumerate(working_days, start=3):
            day_name = days_names[day] if day < len(days_names) else f"Day {day}"
            day_cell = ws.cell(row=row_index, column=1)
            day_cell.fill = day_fill
            apply_header(day_cell, day_name, fg="FFFFFF")
            ws.row_dimensions[row_index].height = 40

            for col_index, period in enumerate(all_periods, start=2):
                matching = [
                    slot
                    for slot in section_slots
                    if slot.get("day") == day
                    and (
                        slot.get("period") == period
                        or (slot.get("slot_type") == "lab" and slot.get("period") == period - 1)
                    )
                ]
                cell = ws.cell(row=row_index, column=col_index)
                if matching:
                    slot = matching[0]
                    if slot.get("slot_type") == "break":
                        text = "Break Lecture\nNo substitute available\nFree period"
                        cell.fill = break_fill
                    else:
                        text = (
                            f"{slot.get('course_name', '')}\n"
                            f"{slot.get('faculty_name', '')}\n"
                            f"{slot.get('room_name', '')}"
                        )
                        cell.fill = lab_fill if slot.get("slot_type") == "lab" else theory_fill
                    cell.value = text
                    cell.font = Font(name="Calibri", size=9)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.fill = empty_fill
                cell.border = cell_border

        ws.column_dimensions["A"].width = 14
        for col_index in range(2, len(all_periods) + 2):
            ws.column_dimensions[get_column_letter(col_index)].width = 18

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pdf(
    slots: List[Dict[str, Any]],
    institution_name: str,
    periods_per_day: Dict,
    days_names: Optional[List[str]] = None,
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    if days_names is None:
        days_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle(
        "cell",
        fontSize=7,
        leading=9,
        alignment=1,
        fontName="Helvetica",
    )
    header_style = ParagraphStyle(
        "hdr",
        fontSize=8,
        leading=10,
        alignment=1,
        fontName="Helvetica-Bold",
        textColor=colors.white,
    )
    title_style = ParagraphStyle(
        "title",
        fontSize=14,
        leading=18,
        alignment=1,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#1E293B"),
    )

    section_entries = _section_entries(slots)
    working_days = sorted(periods_per_day.keys())
    all_periods = sorted({period for periods in periods_per_day.values() for period in periods})
    story = []

    navy = colors.HexColor("#1E293B")
    slate = colors.HexColor("#334155")
    green = colors.HexColor("#D1FAE5")
    amber = colors.HexColor("#FEF3C7")
    white = colors.white

    for section_id, section_name in section_entries:
        section_slots = _slots_for_section(slots, section_id)

        story.append(Paragraph(f"{institution_name}", title_style))
        story.append(Paragraph(f"Section: {section_name}", styles["Heading2"]))
        story.append(Spacer(1, 4 * mm))

        header_row = [Paragraph("Day / Period", header_style)] + [
            Paragraph(f"P{period + 1}", header_style)
            for period in all_periods
        ]
        table_data = [header_row]

        for day in working_days:
            day_name = days_names[day] if day < len(days_names) else f"Day {day}"
            row = [
                Paragraph(
                    day_name,
                    ParagraphStyle(
                        "day",
                        fontSize=8,
                        fontName="Helvetica-Bold",
                        alignment=1,
                        textColor=colors.white,
                    ),
                )
            ]
            for period in all_periods:
                matching = [
                    slot
                    for slot in section_slots
                    if slot.get("day") == day
                    and (
                        slot.get("period") == period
                        or (slot.get("slot_type") == "lab" and slot.get("period") == period - 1)
                    )
                ]
                if matching:
                    slot = matching[0]
                    if slot.get("slot_type") == "break":
                        text = "<b>Break Lecture</b><br/>No substitute available<br/><i>Free period</i>"
                    else:
                        text = (
                            f"<b>{slot.get('course_name', '')}</b><br/>"
                            f"{slot.get('faculty_name', '')}<br/>"
                            f"<i>{slot.get('room_name', '')}</i>"
                        )
                    row.append(Paragraph(text, cell_style))
                else:
                    row.append(Paragraph("", cell_style))
            table_data.append(row)

        col_widths = [30 * mm] + [22 * mm] * len(all_periods)
        table = Table(table_data, colWidths=col_widths, repeatRows=1)

        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), navy),
            ("BACKGROUND", (0, 1), (0, -1), slate),
            ("TEXTCOLOR", (0, 0), (-1, 0), white),
            ("TEXTCOLOR", (0, 1), (0, -1), white),
            ("ROWBACKGROUNDS", (1, 1), (-1, -1), [white, colors.HexColor("#F1F5F9")]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ROWHEIGHT", (1, 0), (-1, -1), 26),
        ]

        for row_index, day in enumerate(working_days, start=1):
            for col_index, period in enumerate(all_periods, start=1):
                matching = [
                    slot
                    for slot in section_slots
                    if slot.get("day") == day
                    and (
                        slot.get("period") == period
                        or (slot.get("slot_type") == "lab" and slot.get("period") == period - 1)
                    )
                ]
                if matching:
                    fill_color = amber if matching[0].get("slot_type") == "break" else green
                    style_cmds.append(("BACKGROUND", (col_index, row_index), (col_index, row_index), fill_color))

        table.setStyle(TableStyle(style_cmds))
        story.append(table)
        story.append(PageBreak())

    if story:
        story = story[:-1]
    doc.build(story)
    return buf.getvalue()
