import contextlib
import io
import os
import sys
import unittest
from pathlib import Path


BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

os.environ.setdefault("DATABASE_URL", "sqlite:///./test_timetable.db")

from fastapi.testclient import TestClient

from database import SessionLocal, init_db
from main import app
from models import Course, Department, Faculty, Section, SectionCourse
from seed import seed


class BackendTestCase(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        init_db()
        cls.client = TestClient(app)

    def setUp(self):
        with contextlib.redirect_stdout(io.StringIO()):
            seed()

    def _generate(
        self,
        institution_id: int = 1,
        name: str = "Test Timetable",
        department_id: int | None = None,
        semester: int | None = None,
        expected_status_code: int = 200,
    ):
        payload: dict = {
            "institution_id": institution_id,
            "name": name,
            "max_solve_seconds": 20,
        }
        if department_id is not None:
            payload["department_id"] = department_id
        if semester is not None:
            payload["semester"] = semester

        response = self.client.post("/timetables/generate", json=payload)
        self.assertEqual(response.status_code, expected_status_code, response.text)
        return response

    def _add_semester_six_scope(self):
        db = SessionLocal()
        dept = db.query(Department).filter(Department.name == "Computer Science").first()
        faculty = db.query(Faculty).filter(Faculty.name == "Dr. Anita Sharma").first()
        course = db.query(Course).filter(Course.name == "Data Structures").first()

        section = Section(
            department_id=dept.id,
            name="CS-E",
            student_count=45,
            semester=6,
            class_representative_name="Riya",
            class_representative_phone="+911234567890",
        )
        db.add(section)
        db.commit()
        db.refresh(section)

        db.add(SectionCourse(section_id=section.id, course_id=course.id, faculty_id=faculty.id))
        db.commit()

        result = {
            "department_id": dept.id,
            "section_id": section.id,
            "faculty_id": faculty.id,
        }
        db.close()
        return result

    def test_generate_seed_data_success_and_no_overlaps(self):
        generated = self._generate(name="Seed Timetable").json()
        self.assertEqual(generated["status"], "done")
        self.assertEqual(generated["conflicts"], [])

        detail = self.client.get(f"/timetables/{generated['timetable_id']}")
        self.assertEqual(detail.status_code, 200, detail.text)
        slots = detail.json()["slots"]

        faculty_usage = set()
        section_usage = set()
        room_usage = set()
        for slot in slots:
            occupied_periods = [slot["period"]]
            if slot["duration"] > 1:
                occupied_periods.append(slot["period"] + 1)

            for period in occupied_periods:
                faculty_key = (slot["faculty_id"], slot["day"], period)
                self.assertNotIn(faculty_key, faculty_usage)
                faculty_usage.add(faculty_key)

                for section_id in slot.get("section_ids") or ([slot["section_id"]] if slot["section_id"] else []):
                    section_key = (section_id, slot["day"], period)
                    self.assertNotIn(section_key, section_usage)
                    section_usage.add(section_key)

                if slot["room_id"] is not None:
                    room_key = (slot["room_id"], slot["day"], period)
                    self.assertNotIn(room_key, room_usage)
                    room_usage.add(room_key)

    def test_generation_requires_explicit_semester_when_scope_is_ambiguous(self):
        self._add_semester_six_scope()

        response = self._generate(
            name="Ambiguous Timetable",
            expected_status_code=400,
        )
        self.assertIn("Select a semester", response.json()["detail"])

    def test_sections_allow_same_name_in_different_semesters_but_not_same_semester(self):
        allowed = self.client.post(
            "/sections",
            json={
                "department_id": 1,
                "name": "CS-A",
                "student_count": 48,
                "semester": 6,
            },
        )
        self.assertEqual(allowed.status_code, 200, allowed.text)
        self.assertEqual(allowed.json()["name"], "CS-A")
        self.assertEqual(allowed.json()["semester"], 6)

        blocked = self.client.post(
            "/sections",
            json={
                "department_id": 1,
                "name": "CS-A",
                "student_count": 48,
                "semester": 5,
            },
        )
        self.assertEqual(blocked.status_code, 400, blocked.text)
        self.assertIn("selected semester", blocked.json()["detail"])

    def test_scoped_generation_only_uses_selected_semester_and_views_pick_matching_timetable(self):
        extra_scope = self._add_semester_six_scope()

        sem_five = self._generate(
            name="Computer Science Sem 5",
            department_id=extra_scope["department_id"],
            semester=5,
        ).json()
        sem_six = self._generate(
            name="Computer Science Sem 6",
            department_id=extra_scope["department_id"],
            semester=6,
        ).json()

        sem_five_detail = self.client.get(f"/timetables/{sem_five['timetable_id']}").json()
        sem_five_section_names = {slot["section_name"] for slot in sem_five_detail["slots"] if slot.get("section_name")}
        self.assertIn("CS-A", sem_five_section_names)
        self.assertNotIn("CS-E", sem_five_section_names)

        timetable_list = self.client.get("/timetables", params={"institution_id": 1})
        self.assertEqual(timetable_list.status_code, 200, timetable_list.text)
        listed = timetable_list.json()
        sem_six_meta = next(item for item in listed if item["id"] == sem_six["timetable_id"])
        self.assertEqual(sem_six_meta["department_name"], "Computer Science")
        self.assertEqual(sem_six_meta["semester_number"], 6)
        self.assertIn("Semester 6", sem_six_meta["scope_label"])

        student_view = self.client.get(
            "/views/student",
            params={
                "institution_id": 1,
                "department_id": extra_scope["department_id"],
                "semester": 6,
                "section_id": extra_scope["section_id"],
            },
        )
        self.assertEqual(student_view.status_code, 200, student_view.text)
        student_payload = student_view.json()
        self.assertEqual(student_payload["timetable_id"], sem_six["timetable_id"])
        self.assertEqual(student_payload["semester"], 6)
        self.assertTrue(all(slot["section_name"] == "CS-E" for slot in student_payload["slots"]))

        teacher_options = self.client.get(
            "/views/teacher/faculty",
            params={
                "institution_id": 1,
                "department_id": extra_scope["department_id"],
                "semester": 6,
            },
        )
        self.assertEqual(teacher_options.status_code, 200, teacher_options.text)
        teacher_ids = {item["id"] for item in teacher_options.json()}
        self.assertIn(extra_scope["faculty_id"], teacher_ids)

        teacher_view = self.client.get(
            "/views/teacher",
            params={
                "institution_id": 1,
                "department_id": extra_scope["department_id"],
                "semester": 6,
                "faculty_id": extra_scope["faculty_id"],
            },
        )
        self.assertEqual(teacher_view.status_code, 200, teacher_view.text)
        teacher_payload = teacher_view.json()
        self.assertEqual(teacher_payload["timetable_id"], sem_six["timetable_id"])
        self.assertEqual(teacher_payload["semester"], 6)
        self.assertTrue(any(slot["section_name"] == "CS-E" for slot in teacher_payload["slots"]))

    def test_analytics_and_exports_work_for_scoped_timetable(self):
        generated = self._generate(name="Analytics Timetable").json()

        analytics = self.client.get(f"/timetables/{generated['timetable_id']}/analytics")
        self.assertEqual(analytics.status_code, 200, analytics.text)
        analytics_payload = analytics.json()
        self.assertIn("faculty_load", analytics_payload)
        self.assertIn("room_utilization", analytics_payload)
        self.assertIn("core_subject_distribution", analytics_payload)

        excel = self.client.get(f"/timetables/{generated['timetable_id']}/export/excel")
        pdf = self.client.get(f"/timetables/{generated['timetable_id']}/export/pdf")
        self.assertEqual(excel.status_code, 200, excel.text)
        self.assertEqual(pdf.status_code, 200, pdf.text)
        self.assertGreater(len(excel.content), 100)
        self.assertGreater(len(pdf.content), 100)

    def test_combined_group_is_visible_for_all_sections_and_in_exports(self):
        generated = self._generate(name="Combined Visibility").json()
        detail = self.client.get(f"/timetables/{generated['timetable_id']}").json()
        combined_slot = next(slot for slot in detail["slots"] if slot["is_combined"])
        second_section_id = combined_slot["section_ids"][1]

        student_view = self.client.get(
            "/views/student",
            params={
                "institution_id": 1,
                "department_id": 1,
                "semester": 5,
                "section_id": second_section_id,
            },
        )
        self.assertEqual(student_view.status_code, 200, student_view.text)
        student_slots = student_view.json()["slots"]
        self.assertTrue(
            any(
                slot["course_id"] == combined_slot["course_id"]
                and slot["day"] == combined_slot["day"]
                and slot["period"] == combined_slot["period"]
                and second_section_id in (slot.get("section_ids") or [])
                for slot in student_slots
            )
        )

        excel = self.client.get(f"/timetables/{generated['timetable_id']}/export/excel")
        self.assertEqual(excel.status_code, 200, excel.text)
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(excel.content))
        self.assertIn("CS-B", workbook.sheetnames)
        cs_b_values = [
            cell.value
            for row in workbook["CS-B"].iter_rows()
            for cell in row
            if isinstance(cell.value, str)
        ]
        self.assertTrue(any("Engineering Maths" in value for value in cs_b_values))

    def test_generation_can_reuse_locked_slots_from_source_timetable(self):
        base = self._generate(name="Locked Source", department_id=1, semester=5).json()
        detail = self.client.get(f"/timetables/{base['timetable_id']}").json()
        locked_candidates = [slot for slot in detail["slots"] if slot["slot_type"] != "break"][:2]

        for slot in locked_candidates:
            response = self.client.patch(f"/slots/{slot['id']}/lock")
            self.assertEqual(response.status_code, 200, response.text)

        regenerated = self.client.post(
            "/timetables/generate",
            json={
                "institution_id": 1,
                "department_id": 1,
                "semester": 5,
                "name": "Locked Regenerated",
                "source_timetable_id": base["timetable_id"],
                "max_solve_seconds": 20,
            },
        )
        self.assertEqual(regenerated.status_code, 200, regenerated.text)
        regenerated_payload = regenerated.json()
        self.assertEqual(regenerated_payload["locked_slots_used"], len(locked_candidates))

        regenerated_detail = self.client.get(f"/timetables/{regenerated_payload['timetable_id']}").json()
        for locked in locked_candidates:
            self.assertTrue(
                any(
                    slot["day"] == locked["day"]
                    and slot["period"] == locked["period"]
                    and slot["course_id"] == locked["course_id"]
                    and slot["faculty_id"] == locked["faculty_id"]
                    and sorted(slot.get("section_ids") or []) == sorted(locked.get("section_ids") or [])
                    for slot in regenerated_detail["slots"]
                )
            )

    def test_sections_persist_cr_details_and_nlp_execute_updates_faculty_rules(self):
        section_response = self.client.post(
            "/sections",
            json={
                "department_id": 1,
                "name": "CS-Z",
                "student_count": 52,
                "semester": 7,
                "class_representative_name": "Asha",
                "class_representative_phone": "+919999999999",
            },
        )
        self.assertEqual(section_response.status_code, 200, section_response.text)
        section_payload = section_response.json()
        self.assertEqual(section_payload["class_representative_name"], "Asha")
        self.assertEqual(section_payload["class_representative_phone"], "+919999999999")

        listed_sections = self.client.get("/sections", params={"department_id": 1})
        self.assertEqual(listed_sections.status_code, 200, listed_sections.text)
        listed_payload = next(section for section in listed_sections.json() if section["id"] == section_payload["id"])
        self.assertEqual(listed_payload["class_representative_name"], "Asha")
        self.assertEqual(listed_payload["class_representative_phone"], "+919999999999")

        unavailability = self.client.post(
            "/nlp/execute",
            json={
                "institution_id": 1,
                "text": "Dr. Anita Sharma cannot teach before 11am on Monday",
            },
        )
        self.assertEqual(unavailability.status_code, 200, unavailability.text)
        self.assertTrue(unavailability.json()["executed"])

        max_consecutive = self.client.post(
            "/nlp/execute",
            json={
                "institution_id": 1,
                "text": "Dr. Anita Sharma 2 consecutive periods maximum",
            },
        )
        self.assertEqual(max_consecutive.status_code, 200, max_consecutive.text)
        self.assertTrue(max_consecutive.json()["executed"])

        db = SessionLocal()
        faculty = db.query(Faculty).filter(Faculty.name == "Dr. Anita Sharma").first()
        self.assertIsNotNone(faculty)
        blocked = {(slot["day"], slot["period"]) for slot in (faculty.unavailable_slots or [])}
        self.assertIn((0, 1), blocked)
        self.assertEqual(faculty.max_consecutive_periods, 2)
        db.close()


if __name__ == "__main__":
    unittest.main()
